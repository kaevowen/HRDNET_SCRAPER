import os
import re
import json

from datetime import datetime
from urllib.parse import quote

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Side, Border
from requests import get, Session
from bs4 import BeautifulSoup as BS

RE_COMP = re.compile('[b\\/:*?"<>|\\t]')
EXCEL_BASE_URL = "http://www.hrd.go.kr/hrdp/ps/ppsmo/excelDownAll0109P.do?"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DL_DIR = os.path.join(BASE_DIR, "result")
HEADERS = {
    "Accept": "text/html,application/xhtml+xml,"
              "application/xml;q=0.9,image/avif,"
              "image/webp,image/apng,*/*;q=0.8,"
              "application/signed-exchange;v=b3;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "DNT": "1",
    "Host": "www.hrd.go.kr",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/86.0.4240.198 Safari/537.36"
}
if not os.path.isdir(DL_DIR):
    os.mkdir(DL_DIR)


def getExcel(sess, title, subtitle, tracseId):
    args = f"tracseId={tracseId}&ncsYn=Y&pssrpYear={datetime.now().year}&pssrpTme=9&mainTracseSe=&ncsAbluitFactorUnitSe=&tracseSttusCd=&jdgmnSe=&excelAllYn=&excelGbn=&A2Gbn="

    t = sess.get(EXCEL_BASE_URL + args)
    if t.headers.get('content-type') == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        print("Downloading... ", end=' ')
        FILE_DIR = os.path.join(DL_DIR, f"{title}.xlsx")
        with open(f'{FILE_DIR}', 'wb') as f:
            f.write(t.content)

        wb = openpyxl.load_workbook(FILE_DIR)
        ws = wb[wb.sheetnames[0]]
        set_template(ws, subtitle)
        wb.copy_worksheet(wb[wb.sheetnames[0]])
        wb[wb.sheetnames[1]].delete_cols(12)
        wb.save(os.path.join(DL_DIR, f'{title}_{subtitle}.xlsx'))
        os.remove(FILE_DIR)


def set_template(ws, acaName):
    max_rows = 0
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    align = Alignment(horizontal='center', vertical='center', wrapText=True)

    ws.merge_cells('A1:L1')
    ws.merge_cells('A2:L2')
    ws['A1'] = '학원명'
    ws['A2'] = acaName
    ws['A1'].fill = PatternFill(start_color='FFFF00',
                                end_color='FFFF00',
                                fill_type='solid')

    for _ in ws.rows:
        max_rows += 1

    rows = ws.iter_rows(1, max_rows)
    for row in rows:
        for cell in row:
            cell.border = border
            cell.alignment = align

    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF00',
                                    end_color='FFFF00',
                                    fill_type='solid')


def checkLogin(ID, PW):
    s = Session()
    s.headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko'
    }
    res = s.get(
        'https://www.hrd.go.kr/hrdp/mb/pmbao/login.do?'
        f'userloginId={ID}&'
        f'userloginPwd={PW}&'
        'loginType=personal&'
        'mberSe=C0178'
    )

    res = json.loads(res.text)
    if res['message'] == 'checkLoginId' or res['message'] == 'FAIL':
        return "Fail_Login"


    if not checkAuthKey(s):
        return "Fail_Authkey"
    
    return s


def checkAuthKey(s):
    ak = os.path.join(BASE_DIR, "authKey.key")
    if not os.path.isfile(ak):
        h = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6',
            'Connection': 'keep-alive',
            'Cookie': 'WMONID=sY_dRJFIDLl; JSESSIONID=nY2JggPZNpl2f52Lv5K31Dcy0bQcMCbTRtp6PH6C5FNXmvvTvQ21!673857750!-761163285; PCID=16127466537135153594340; RC_RESOLUTION=1600*900; RC_COLOR=24; openpop5=; UID=kaev6',
            'DNT': '1',
            'Host': 'www.hrd.go.kr',
            'Referer': 'https://www.hrd.go.kr/hrdp/ps/ppsho/PPSHO0100L.do',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.146 Safari/537.36'
        }
        res = BS(s.get('https://www.hrd.go.kr/hrdp/ps/ppsho/PPSHO0100L.do').content, "html.parser")
        try:
            key = re.sub('\s', '', res.select('table.view > tbody > tr > td')[-1].text)
            
            with open(ak, "w+") as f:
                f.write(key)
                
        except IndexError:
            return False
        
    return True
    
class HrdNetAPI:
    def __init__(self, sess, startDate, endDate, flag, NcsCode, AreaCode, keyword=None, keyword2=None):
        with open('authKey.key') as f:
            self.authKey = f.readline()

        self.startDate = startDate
        self.endDate = endDate
        self.flag = flag
        self.keyword = keyword
        self.keyword2 = keyword2
        self.NcsCode = NcsCode
        self.AreaCode = AreaCode
        self.pagination = 0
        self.cnt = 0
        self.ContentCnt = 0
        self.session = sess
        self.url = f'http://www.hrd.go.kr/jsp/HRDP/HRDPO00/HRDPOA{self.flag}/HRDPOA{self.flag}_1.jsp?'
        self.payload = ''
        self.urlWithPayload = ''

    def getPagination(self):
        self.payload = f'returnType=XML&' \
                  f'authKey={self.authKey}&' \
                  f'pageNum=1&' \
                  f'pageSize=100&' \
                  f'srchTraStDt={self.startDate}&' \
                  f'srchTraEndDt={self.endDate}&' \
                  f'outType=1&' \
                  f'sort=ASC&' \
                  f'sortCol=TR_STT_DT&'

        if self.keyword is not None:
            k = quote(self.keyword)
            self.payload += f'srchTraProcessNm={k}&'

        if self.keyword2 is not None:
            k = quote(self.keyword2)
            self.payload += f'srchTraOrganNm={k}&'

        for i in range(len(self.NcsCode)):
            if self.NcsCode[i] is not None:
                self.payload += f'srchKeco{ i + 1 }={self.NcsCode[i]}&'

        for i in range(len(self.AreaCode)):
            if self.AreaCode[i] is not None:
                self.payload += f'srchTraArea{ i + 1 }={self.AreaCode[i]}&'

        self.urlWithPayload = self.url + self.payload
        res = get(self.urlWithPayload, headers=HEADERS)
        tmp = BS(res.content, "lxml")
        self.ContentCnt = int(tmp.find('scn_cnt').text)
        self.pagination = (self.ContentCnt // 100) + 2
        return self.ContentCnt

    def getAPI(self):
        print(f"total : {self.ContentCnt}")

        for i in range(1, self.pagination):
            tmp = re.sub('pageNum=\\d{1,2}', f'pageNum={i}', self.urlWithPayload)
            content = BS(get(tmp, headers=HEADERS).content, "lxml")

            print(tmp)

            for e in content.select('scn_list'):
                    tracseId = e.select("trprid")[0].text
                    title = re.sub(RE_COMP, "", e.select('title')[0].text)
                    subtitle = re.sub(RE_COMP, "", e.select('subtitle')[0].text)
                    getExcel(self.session, title, subtitle, tracseId)
                    self.cnt += 1
                    percentage = round((self.cnt / self.ContentCnt) * 100, 2)
                    print(f"{percentage} %")

        os.startfile(DL_DIR)
